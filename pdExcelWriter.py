# -*- coding: utf-8 -*-
"""
Created on Tue Dec 15 15:23:06 2020

@author: Administrator
"""
import pandas as pd

from openpyxl import load_workbook
import os
import timeit
import datetime
import traceback
class pdExcelWriter:
    def __init__(self, filename):
        self.wb = wb = load_workbook(filename)



    def df2Sheet(self,sheetname,df):        
        sheet = self.wb[sheetname]
        for col in range(1,df.shape[1]+1):
            try:
                
                if df.columns[col-1].find('时间')!=-1:
                    df[df.columns[col-1]]=pd.to_datetime(df[df.columns[col-1]],errors='coerce')                    
            except Exception as e:                
                print(traceback.print_exc())  
                print("df.columns[col-1]="+str(df.columns[col-1]))
            sheet.cell(row=1,column=col,value=df.columns[col-1])                  
            for row in range(1,df.shape[0]+1):
                if df.iloc[row-1,col-1] is not pd.NaT:
                    sheet.cell(row=row+1,column=col,value=df.iloc[row-1,col-1])



    def saveWorkbook(self,filename):
        self.wb.save(filename)
    
    

if __name__ == '__main__':
    starttime = datetime.datetime.now()
    print('开始创建pdew')
    pdew=pdExcelWriter(os.path.join('templates','倍增模板.xlsx'))
    endtime = datetime.datetime.now()
    duringtime = endtime -  starttime        
    starttime=endtime
    print('完成创建pdew for ' + str(duringtime.seconds) + '秒')
    print('开始读入test1.csv')
    
    df=pd.read_csv(os.path.join('templates','test1.csv'),engine='python')
    #df1=pd.read_csv(os.path.join('templates','正常在途数.csv'),engine='python')
    #df2=pd.read_csv(os.path.join('templates','退单在途工单.csv'),engine='python')
    #df['派单时间']=pd.to_datetime(df['派单时间'],errors='coerce')
    #df['归档时间']=pd.to_datetime(df['归档时间'],errors='coerce')
    #df1['派单时间']=pd.to_datetime(df1['派单时间'],errors='coerce')
    #df2['派单时间']=pd.to_datetime(df2['派单时间'],errors='coerce')
    endtime = datetime.datetime.now()
    duringtime = endtime -  starttime   
    starttime=endtime     
    print('完成读入for '+str(duringtime.seconds)+'秒')
    print('开始df2sheet')
    pdew.df2Sheet('市场兑换率（归档数）',df)
    #pdew.df2Sheet('正常在途数',df1)
    #pdew.df2Sheet('退单在途工单',df2)
    endtime = datetime.datetime.now()
    duringtime = endtime -  starttime        
    starttime=endtime
    print('完成 for ' + str(duringtime.seconds)+ '秒')
    print('开始saveworkbook')
    pdew.saveWorkbook(os.path.join('templates','result.xlsx'))
    endtime = datetime.datetime.now()
    duringtime = endtime -  starttime        

    print('完成 for ' + str(duringtime.seconds)+ '秒')
    
    
